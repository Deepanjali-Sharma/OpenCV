#GUI layout of Face detection
#!/usr/bin/env python3
# importing tkinter and tkinter.ttk
# and all their functions and classes
# importing askopenfile function
# from class filedialog
import time
from tkinter import *
from tkinter.ttk import *
from PIL import Image, ImageTk
from tkinter.filedialog import askopenfile

root = Tk()
# show no frame
root.overrideredirect(True)
width = root.winfo_screenwidth()
height = root.winfo_screenheight()
root.geometry('%dx%d+%d+%d' % (width*0.6, height*0.6, width*0.2, height*0.2))
# take a .jpg picture you like, add text with a program like PhotoFiltre
# (free from http://www.photofiltre.com) and save as a .gif image file
image_file = "v199.jpeg"
#assert os.path.exists(image_file)
# use Tkinter's PhotoImage for .gif files
image = ImageTk.PhotoImage(file=image_file)
canvas = Canvas(root, height=height*0.6, width=width*0.6, bg="grey")
canvas.create_image(width*0.6/2, height*0.6/2, image=image)
canvas.pack()
# show the splash screen for 5000 milliseconds then destroy
root.after(5000, root.destroy)
root.mainloop()
time.sleep(1)


root = Tk()

#Creating root of gui
root.title("Image Processing app")
root.geometry('800x400')

# This function will be used to open file in read mode and only image files will be opened
# This function can open any format image in online mode

def open_file():
	file = askopenfile(mode ='r', filetypes =[('Image Files', '*.jpg')])
	print(file)
	if file is not None:
		content = file.name
		print(content)

        # file database of the current file in read mode
		file = open('geek.txt', 'w')
		file.write(content)
		file.close()

def openProgram():

	#importing Computer Vision files
	import cv2
	# importing haar file through cascade classifier
	face_cascade = cv2.CascadeClassifier("haarf1.xml")
	face_cascade1 = cv2.CascadeClassifier("haare2.xml")

	#file database of the current file in read mode
	file = open("geek.txt", "r")
	content = file.read()
	img1 = cv2.imread(content)
	file.close()

	# conversion into grayscale image
	fimg = cv2.cvtColor(img1, cv2.COLOR_BGR2GRAY)

	# using cascade classifier block of ML for face haar file with scale factor of 15 % decrement
	facess = face_cascade.detectMultiScale(fimg, scaleFactor=1.15, minNeighbors=5)

	# using cascade classifier block of ML for eye haar file with scale factor of 10 % decrement
	fac = face_cascade1.detectMultiScale(fimg, scaleFactor=1.13, minNeighbors=20)

	# importing open cv for code vision
	print("face matrix ", facess)

	# importing open cv for code vision
	print("eye matrix ", fac)

	# searching for corners of the face through data
	for x, y, w, h in facess:
		img1 = cv2.rectangle(img1, (x, y), (x + w, y + h), (0, 255, 0), 3)

	# searching for corners of the eyes through data
	for x, y, w, h in fac:
		img1 = cv2.rectangle(img1, (x, y), (x + w, y + h), (0, 255, 0), 3)

	# resizing the image to fit in the running window
	resizedface = cv2.resize(img1, (int(img1.shape[1] / 2), int(img1.shape[0] / 2)))

	# displaying the image with face and eyes block
	cv2.imshow('Gray', resizedface)
	cv2.waitKey(0)
	cv2.destroyWindow("Gray")

	import openpyxl
	path = "demo.xlsx"

	wb_obj = openpyxl.load_workbook(path)
	sheet_obj = wb_obj.active

	maxrow = sheet_obj.max_row
	print(sheet_obj.max_row)
	i= sheet_obj.cell(row=1, column=2)
	print(i.value)
	c1 = sheet_obj.cell(row=maxrow + 1, column=1)
	valueforcell = "##100{}".format(i.value)
	c1.value = valueforcell
	c21 = sheet_obj.cell(row=1, column=2)
	c21.value=i.value+1
	wb_obj.save("demo.xlsx")

	path1 = "demo.xlsx"

	wbp_obj = openpyxl.load_workbook(path1)

	sheetp_obj = wbp_obj.active
	m_row = sheetp_obj.max_row

	# Loop will print all values
	# of first column
	a = []
	for i in range(1, m_row + 1):
		cell_obj = sheetp_obj.cell(row=i, column=1)
		print(cell_obj.value)
		a.append(cell_obj.value)
	print(str('\n'.join(map(str, a))))

	file1 = open("data.txt", "w")
	file1.write(str('\n'.join(map(str, a))))
	file1.close()

	import os
	os.system('data.txt')

	# after key is pressed the image will exit

	# cleaning all windos after button is pressed
#labelling gui contnets that will be shown
w = Label(root, text='  ', font=("Helvetica", 33))
w.pack()

w = Label(root, text='Firstly Press to Select Image', font=("Helvetica", 13))
w.pack()

#labelling gui buttons that will be shown
btn = Button(root, text ='Open', command = lambda:open_file())
btn.pack(side = TOP, pady = 10)

w = Label(root, text='Secondaly Press to Detect', font=("Helvetica", 13))
w.pack()

mbutton = Button(text = "Go", command = openProgram).pack()
btn.pack(side = TOP, pady = 10)

w = Label(root, text='  ', font=("Helvetica", 73))
w.pack()

w = Label(root, text='Image Detection App', font=("Helvetica", 56))
w.pack()

#images in the gui
load = Image.open("v9.jpg")
render = ImageTk.PhotoImage(load)
img = Label(image=render)
img.image = render
img.place(x=0, y=0)

#closing and finishing the gui layout
mainloop()





